#!/usr/bin/env python3
"""
fill_formulario.py - Llena el formulario oficial MSP con datos del paciente
Recibe JSON con datos, llena el Excel, exporta PDF
"""
from openpyxl import load_workbook
from openpyxl.styles.borders import Border
import json, sys, os, subprocess, tempfile

def fill_formulario(data, template_path, output_path, modo='todo'):
    wb = load_workbook(template_path)
    
    def w(ws, cell, val, align='left', font_size=9):
        if val is None or val == '':
            return
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Alignment, Font
        wrap = Alignment(horizontal=align, vertical='center', wrap_text=True)
        fnt = Font(size=font_size)
        target = ws[cell]
        if isinstance(target, MergedCell):
            for m in ws.merged_cells.ranges:
                if cell in m:
                    tl = ws.cell(row=m.min_row, column=m.min_col)
                    tl.value = str(val)
                    tl.alignment = wrap
                    tl.font = fnt
                    return
        else:
            ws[cell] = str(val)
            target = ws[cell]
            target.alignment = wrap
            target.font = fnt
    
    def ck(ws, cell, condition):
        from openpyxl.cell.cell import MergedCell
        target = ws[cell]
        val = 'X' if condition else ''
        if isinstance(target, MergedCell):
            for m in ws.merged_cells.ranges:
                if cell in m:
                    ws.cell(row=m.min_row, column=m.min_col).value = val
                    return
        else:
            ws[cell] = val

    def insertar_firma(ws, firma_b64, col_letra, fila, ancho_px=160, alto_px=60):
        """Inserta imagen base64 en la hoja en la celda indicada"""
        if not firma_b64:
            return
        try:
            import base64
            from openpyxl.drawing.image import Image as XLImage
            from io import BytesIO
            # Quitar prefijo data:image/...;base64,
            if ',' in firma_b64:
                firma_b64 = firma_b64.split(',', 1)[1]
            img_data = base64.b64decode(firma_b64)
            img_stream = BytesIO(img_data)
            img = XLImage(img_stream)
            img.width = ancho_px
            img.height = alto_px
            # Anclar en la celda indicada
            from openpyxl.utils import get_column_letter
            if isinstance(col_letra, int):
                col_letra = get_column_letter(col_letra)
            img.anchor = f'{col_letra}{fila}'
            ws.add_image(img)
        except Exception as e:
            print(f'⚠️ Error insertando firma: {e}')

    # ========================================
    # Datos comunes
    # ========================================
    if 'ficha' in data:
        pac = data.get('paciente', {})
        ec = data.get('empresa', {})
        fo = data.get('ficha', {})
        prof_nombre = data.get('profesional', {}).get('nombre', '')
        prof_codigo = data.get('profesional', {}).get('codigo', '')
    else:
        fo = data
        pac = data.get('_paciente', {})
        ec = data.get('_empresa', {})
        prof_nombre = ''
        prof_codigo = ''

    # Firma desde payload
    firma_b64 = data.get('_firma', '')

    a = fo.get('a', {})
    b = fo.get('b', {})
    c = fo.get('c', {})
    d = fo.get('d', {})
    e = fo.get('e', {})
    f = fo.get('f', {})
    g = fo.get('g', {})
    h_sec = fo.get('h', {})
    i_sec = fo.get('i', {})
    j = fo.get('j', {})
    k = fo.get('k', {})
    l = fo.get('l', {})
    m = fo.get('m', {})
    n = fo.get('n', {})
    r = fo.get('r', {})
    
    ap1 = a.get('ap1', pac.get('apellido1', ''))
    ap2 = a.get('ap2', pac.get('apellido2', ''))
    n1 = a.get('n1', pac.get('nombre1', ''))
    n2 = a.get('n2', pac.get('nombre2', ''))
    cedula = a.get('nform', pac.get('cedula', ''))
    sexo_raw = a.get('sexo', pac.get('sexo', ''))
    sexo = 'Masculino' if sexo_raw in ('M', 'Masculino') else 'Femenino' if sexo_raw in ('F', 'Femenino') else sexo_raw
    sexo_corto = 'M' if sexo == 'Masculino' else 'F' if sexo == 'Femenino' else sexo_raw
    fecha_nac = pac.get('fecha_nacimiento', '')
    edad = pac.get('edad', '')
    
    ruc = a.get('ruc', ec.get('ruc', ''))
    ciiu = a.get('ciiu', ec.get('ciiu', ''))
    empresa_nombre = a.get('empresa', ec.get('nombre', ''))
    centro_trabajo = a.get('centro', empresa_nombre)
    
    if not prof_nombre:
        prof_nombre = n.get('medico', '')
    if not prof_codigo:
        prof_codigo = n.get('codigo', '')
    
    tipo_eval_raw = b.get('tipo_eval', b.get('tipo', ''))
    tipo_map = {'INGRESO':'INGRESO','PERIÓDICO':'PERIODICO','PERIODICO':'PERIODICO','REINTEGRO':'REINTEGRO','RETIRO':'RETIRO'}
    tipo_eval = tipo_map.get(tipo_eval_raw.upper(), tipo_eval_raw.upper()) if tipo_eval_raw else ''
    
    aptitud_raw = l.get('aptitud', '')
    aptitud_map = {'APTO':'APTO','APTO EN OBSERVACIÓN':'APTO_OBS','APTO CON LIMITACIONES':'APTO_LIM','NO APTO':'NO_APTO','APTO_OBS':'APTO_OBS','APTO_LIM':'APTO_LIM','NO_APTO':'NO_APTO'}
    aptitud = aptitud_map.get(aptitud_raw, aptitud_raw)
    
    def parse_fecha(f):
        if not f: return ('','','')
        parts = f.split('-') if '-' in f else f.split('/')
        if len(parts) == 3:
            return (parts[0], parts[1], parts[2])
        return ('','','')
    
    fecha_eval = b.get('fecha_atencion', b.get('fecha', ''))
    f_y, f_m, f_d = parse_fecha(fecha_eval)
    fn_y, fn_m, fn_d = parse_fecha(fecha_nac)
    
    # ========================================
    # HOJA 1: CERTIFICADO
    # ========================================
    ws1 = wb['CERTIFICADO']
    
    w(ws1, 'L4', ruc, 'center')
    w(ws1, 'R4', ciiu, 'center')
    w(ws1, 'V4', empresa_nombre, 'center')
    w(ws1, 'AC4', a.get('nform', cedula), 'center')
    w(ws1, 'AI4', a.get('narch', cedula), 'center')
    
    w(ws1, 'A6', ap1, 'center')
    w(ws1, 'J6', ap2, 'center')
    w(ws1, 'Q6', n1, 'center')
    w(ws1, 'X6', n2, 'center')
    w(ws1, 'AD6', sexo_corto, 'center')
    w(ws1, 'AG6', b.get('puesto_ciuo', ''), 'center')
    ws1.row_dimensions[6].height = 25
    
    from openpyxl.styles import Alignment
    center = Alignment(horizontal='center', vertical='center')
    
    w(ws1, 'K10', f_y, 'center')
    ws1['K10'].alignment = center
    w(ws1, 'M10', f_m[0] if len(f_m) >= 1 else '')
    ws1['M10'].alignment = center
    w(ws1, 'N10', f_m[1] if len(f_m) >= 2 else '')
    ws1['N10'].alignment = center
    w(ws1, 'O10', f_d[0] if len(f_d) >= 1 else '')
    ws1['O10'].alignment = center
    w(ws1, 'P10', f_d[1] if len(f_d) >= 2 else '')
    ws1['P10'].alignment = center
    
    ck(ws1, 'L12', tipo_eval == 'INGRESO')
    ws1['L12'].alignment = center
    ck(ws1, 'U12', tipo_eval == 'PERIODICO')
    ws1['U12'].alignment = center
    ck(ws1, 'AC12', tipo_eval == 'REINTEGRO')
    ws1['AC12'].alignment = center
    ck(ws1, 'AI12', tipo_eval == 'RETIRO')
    ws1['AI12'].alignment = center
    
    ck(ws1, 'I17', aptitud == 'APTO')
    ws1['I17'].alignment = center
    ck(ws1, 'S17', aptitud == 'APTO_OBS')
    ws1['S17'].alignment = center
    ck(ws1, 'AC17', aptitud == 'APTO_LIM')
    ws1['AC17'].alignment = center
    ck(ws1, 'AK17', aptitud == 'NO_APTO')
    ws1['AK17'].alignment = center
    
    obs_text = l.get('observaciones', '')
    w(ws1, 'A19', obs_text, 'center')
    ws1.row_dimensions[19].height = 25
    ws1.row_dimensions[20].height = 25
    
    from openpyxl.styles import Alignment, Font
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    for row in [25, 26, 27]:
        try:
            ws1.unmerge_cells(f'A{row}:AL{row}')
        except:
            pass
    try:
        ws1.merge_cells('A25:AL27')
    except:
        pass
    
    ws1.row_dimensions[25].height = 22
    ws1.row_dimensions[26].height = 22
    ws1.row_dimensions[27].height = 22
    
    recs_list = m.get('estandar', [])
    recs_text = ' - '.join(recs_list)
    if m.get('medicas'):
        recs_text += ' - ' + m.get('medicas', '')
    if not recs_text:
        recs_text = (' - '.join(m.get('estandar',[])) + (' - ' + m.get('medicas','') if m.get('medicas') else '')) or m.get('descripcion','')
    
    ws1['A25'] = recs_text
    ws1['A25'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1['A25'].font = Font(size=10)
    
    w(ws1, 'E33', prof_nombre, 'center')
    w(ws1, 'O33', prof_codigo, 'center')

    # ========================================
    # INSERTAR FIRMA EN CERTIFICADO (hoja 1)
    # Celda de firma: cols W33:AL33 (col 23 = W)
    # ========================================
    if firma_b64 and (modo in ('certificado', 'todo')):
        insertar_firma(ws1, firma_b64, col_letra='W', fila=33, ancho_px=155, alto_px=58)
        print('✅ Firma insertada en CERTIFICADO')

    # ========================================
    # HOJA 2: EVALUACION 1/3
    # ========================================
    ws2 = wb.worksheets[1]
    from openpyxl.styles import Alignment
    center = Alignment(horizontal='center', vertical='center')
    
    w(ws2, 'Q5', ruc, 'center')
    w(ws2, 'Z5', ciiu, 'center')
    w(ws2, 'AC5', empresa_nombre, 'center')
    w(ws2, 'AP5', a.get('nform', cedula), 'center')
    w(ws2, 'BB5', a.get('narch', cedula), 'center')
    
    w(ws2, 'B8', ap1, 'center')
    w(ws2, 'S8', ap2, 'center')
    w(ws2, 'AH8', n1, 'center')
    w(ws2, 'AR8', n2, 'center')
    
    ck(ws2, 'U13', sexo == 'Masculino')
    ws2['U13'].alignment = center
    ck(ws2, 'X13', sexo == 'Femenino')
    ws2['X13'].alignment = center
    
    w(ws2, 'Z13', fn_y, 'center')
    ws2['Z13'].alignment = center
    w(ws2, 'AC13', fn_m, 'center')
    ws2['AC13'].alignment = center
    w(ws2, 'AE13', fn_d, 'center')
    ws2['AE13'].alignment = center
    w(ws2, 'AF13', edad, 'center')
    ws2['AF13'].alignment = center
    w(ws2, 'AH13', a.get('grupo_sang', ''), 'center')
    
    ap = a.get('atencion_prioritaria', [])
    if 'Embarazada' in ap: ck(ws2, 'B13', True)
    if 'Persona con discapacidad' in ap: ck(ws2, 'G13', True)
    if 'Enfermedad catastrófica' in ap: ck(ws2, 'K13', True)
    if 'Lactancia' in ap: ck(ws2, 'N13', True)
    if 'Adulto mayor' in ap: ck(ws2, 'R13', True)
    
    sexo = a.get('sexo', '')
    if sexo == 'M': ck(ws2, 'U13', True)
    elif sexo == 'F': ck(ws2, 'X13', True)
    
    lat_raw = a.get('lateralidad', '')
    lat_text = 'Derecha' if lat_raw == 'Diestro' else 'Izquierda' if lat_raw == 'Zurdo' else lat_raw
    w(ws2, 'AP13', lat_text, 'center')
    
    w(ws2, 'J16', b.get('puesto_ciuo', ''), 'center')
    w(ws2, 'AH16', b.get('fecha_atencion', ''), 'center')
    fi_fecha = b.get('fecha_ingreso', '')
    w(ws2, 'B17', fi_fecha, 'center')
    w(ws2, 'AD17', b.get('fecha_reintegro', ''), 'center')
    w(ws2, 'AM17', b.get('fecha_ultimo_dia', ''), 'center')
    
    ck(ws2, 'G19', tipo_eval == 'INGRESO')
    ws2['G19'].alignment = center
    ck(ws2, 'Z19', tipo_eval == 'PERIODICO')
    ws2['Z19'].alignment = center
    ck(ws2, 'AI19', tipo_eval == 'REINTEGRO')
    ws2['AI19'].alignment = center
    ck(ws2, 'BA19', tipo_eval == 'RETIRO')
    ws2['BA19'].alignment = center
    
    w(ws2, 'B21', b.get('observacion', ''), 'center')
    
    ant_clin = c.get('ant_clinicos', c.get('clinicos', '')).replace('\n', ' - ')
    ant_fam = c.get('ant_familiares', c.get('familiares', '')).replace('\n', ' - ')
    w(ws2, 'B24', ant_clin)
    ws2.row_dimensions[24].height = 35
    w(ws2, 'B26', ant_fam)
    ws2.row_dimensions[26].height = 35
    
    if c.get('aut_transfusion') == 'Sí': ck(ws2, 'O29', True)
    if c.get('aut_transfusion') == 'No': ck(ws2, 'S29', True)
    if c.get('trat_hormonal') == 'Sí':
        ck(ws2, 'AH29', True)
        w(ws2, 'AL29', c.get('trat_hormonal_detalle', ''))
    if c.get('trat_hormonal') == 'No': ck(ws2, 'BE29', True)
    
    gineco = c.get('gineco', {})
    if gineco:
        w(ws2, 'B33', gineco.get('fum', ''), 'center')
        w(ws2, 'X33', gineco.get('gestas', ''), 'center')
        w(ws2, 'AB33', gineco.get('partos', ''), 'center')
        w(ws2, 'AE33', gineco.get('cesareas', ''), 'center')
        w(ws2, 'AH33', gineco.get('abortos', ''), 'center')
        if gineco.get('planif_si'): ck(ws2, 'AL33', True)
        if gineco.get('planif_cual'): w(ws2, 'AN33', gineco.get('planif_cual', ''))
        if gineco.get('planif_no'): ck(ws2, 'AV33', True)
        if gineco.get('planif_nr'): ck(ws2, 'AZ33', True)
        examg = gineco.get('examenes', [])
        for ei, ex in enumerate(examg[:2]):
            row = 35 + ei
            w(ws2, f'B{row}', ex.get('nombre', ''))
            w(ws2, f'L{row}', ex.get('tiempo', ''), 'center')
            w(ws2, f'U{row}', ex.get('resultado', ''))
    
    masc = c.get('masculino', {})
    if masc:
        if masc.get('planif_si'): ck(ws2, 'AE40', True)
        if masc.get('planif_cual'): w(ws2, 'AF40', masc.get('planif_cual', ''))
        if masc.get('planif_no'): ck(ws2, 'AP40', True)
        if masc.get('planif_nr'): ck(ws2, 'AY40', True)
        examm = masc.get('examenes', [])
        for ei, ex in enumerate(examm[:2]):
            row = 40 + ei
            w(ws2, f'B{row}', ex.get('nombre', ''))
            w(ws2, f'O{row}', ex.get('tiempo', ''), 'center')
    
    sustancias = c.get('sustancias', [])
    sust_rows = {'TABACO': 44, 'ALCOHOL': 45, 'OTRAS': 46}
    for sust in sustancias:
        nombre = (sust.get('nombre', '') or '').upper()
        row = sust_rows.get(nombre, 46)
        if sust.get('consumo'): w(ws2, f'N{row}', sust.get('consumo', ''))
        if sust.get('estado') == 'ex': ck(ws2, f'S{row}', True)
        if sust.get('abstinencia'): w(ws2, f'X{row}', sust.get('abstinencia', ''))
        if sust.get('estado') == 'no': ck(ws2, f'AB{row}', True)
    
    estilo = c.get('estilo_vida', {})
    if estilo:
        for row in [44, 45, 46]:
            if estilo.get('actividad'): w(ws2, f'AI{row}', estilo.get('actividad', ''))
            if estilo.get('tiempo'): w(ws2, f'AL{row}', estilo.get('tiempo', ''))
            break
    
    medicacion = c.get('medicacion', '')
    if medicacion:
        w(ws2, 'AU44', medicacion)
    
    w(ws2, 'B47', c.get('observacion_sustancias', ''))
    w(ws2, 'B51', d.get('descripcion', ''))
    ws2.row_dimensions[51].height = 35
    
    w(ws2, 'B55', e.get('temp', e.get('temperatura', '')), 'center')
    pa_val = str(e.get('pa_s',''))+('/'+str(e.get('pa_d','')) if e.get('pa_d') else '') if e.get('pa_s') else e.get('presion','')
    w(ws2, 'I55', pa_val, 'center')
    w(ws2, 'P55', e.get('fc', ''), 'center')
    w(ws2, 'W55', e.get('fr', ''), 'center')
    w(ws2, 'AC55', e.get('sat', e.get('sat_o2', '')), 'center')
    w(ws2, 'AG55', e.get('peso', ''), 'center')
    w(ws2, 'AI55', e.get('talla', ''), 'center')
    w(ws2, 'AM55', e.get('imc', ''), 'center')
    w(ws2, 'AS55', e.get('perim_abd', e.get('perimetro', '')), 'center')
    
    hallazgos = f.get('hallazgos', {})
    hall_cell_map = {
        '1a': 'H59', '1b': 'H60', '1c': 'H61',
        '2a': 'H63', '2b': 'H64', '2c': 'H65', '2d': 'H66', '2e': 'H68',
        '3a': 'T59', '3b': 'T60', '3c': 'T62',
        '4a': 'T63', '4b': 'T64', '4c': 'T65', '4d': 'T66', '4e': 'T68',
        '5a': 'AC59', '5b': 'AC60', '5c': 'AC62', '5d': 'AC63',
        '6a': 'AC64', '6b': 'AC65',
        '7a': 'AC66', '7b': 'AC69',
        '8a': 'AJ59', '8b': 'AJ60', '8c': 'AJ61',
        '9a': 'AJ62', '9b': 'AJ63',
        '10a': 'AJ64', '10b': 'AJ66', '10c': 'AJ68',
        '11a': 'BF59', '11b': 'BF60',
        '12a': 'BF62', '12b': 'BF63', '12c': 'BF64',
        '13a': 'BF65', '13b': 'BF66', '13c': 'BF67', '13d': 'BF68',
    }
    for key, cell in hall_cell_map.items():
        if hallazgos.get(key):
            ck(ws2, cell, True)
    
    w(ws2, 'B71', f.get('observaciones', f.get('observacion', '')))
    ws2.row_dimensions[71].height = 35
    ws2.row_dimensions[72].height = 35
    
    # ========================================
    # HOJA 3: FACTORES DE RIESGO (2/3)
    # ========================================
    ws3 = wb.worksheets[2]
    
    w(ws3, 'G2', b.get('puesto_ciuo', ''), 'center', 10)
    ws3.row_dimensions[2].height = 25
    
    riesgos = g.get('riesgos', {})
    h_col = {'h1':'G','h2':'I','h3':'K','h4':'M','h5':'N','h6':'O','h7':'P'}
    riesgo_filas = {
        'fisicos_temperaturasaltas': 6, 'fisicos_temperaturasbajas': 7,
        'fisicos_radiacinionizante': 8, 'fisicos_radiacinnoionizante': 9,
        'fisicos_ruido': 10, 'fisicos_vibracin': 11,
        'fisicos_iluminacin': 12, 'fisicos_ventilacin': 13,
        'fisicos_fluidoelctrico': 14,
        'seguridad_faltadesealizacinaseodesorden': 16,
        'seguridad_atrapamientoentremquinasyosuperficies': 18,
        'seguridad_atrapamientoentreojetos': 18,
        'seguridad_atrapamientoentreobetos': 18,
        'seguridad_cadadeojetos': 19, 'seguridad_cadadeobetos': 19,
        'seguridad_cadasalmismonivel': 20,
        'seguridad_cadasadiferentenivel': 21,
        'seguridad_pinchazos': 22, 'seguridad_cortes': 23,
        'seguridad_choquescolisinvehicular': 24,
        'seguridad_atropellamientosporvehculos': 25,
        'seguridad_proyeccindefluidos': 26,
        'seguridad_proyeccindepartculasfragmentos': 27,
        'seguridad_contactoconsuperficiesdetrabajo': 28,
        'seguridad_contactoelctrico': 29,
        'quimicos_polvos': 31, 'quimicos_slidos': 32, 'quimicos_solidos': 32,
        'quimicos_humos': 33, 'quimicos_lquidos': 34, 'quimicos_liquidos': 34,
        'quimicos_vapores': 35, 'quimicos_aerosoles': 36,
        'quimicos_neblinas': 37, 'quimicos_gaseosos': 38,
        'biologicos_virus': 40, 'biologicos_hongos': 41,
        'biologicos_bacterias': 42, 'biologicos_parsitos': 43, 'biologicos_parasitos': 43,
        'biologicos_exposicinavectores': 44, 'biologicos_exposicionavectores': 44,
        'biologicos_exposicinaanimalesselvaticos': 45, 'biologicos_exposicionaanimalesselvaticos': 45,
        'ergonomicos_manejomanualdecargas': 47,
        'ergonomicos_movimientosrepetitivos': 48, 'ergonomicos_movimientorepetitivos': 48,
        'ergonomicos_postursforzadas': 49, 'ergonomicos_posturasforzadas': 49,
        'ergonomicos_trabajosconpvd': 50,
        'ergonomicos_diseoinadecuadodelpuesto': 51, 'ergonomicos_disenoinadecuadodelpuesto': 51,
        'psicosociales_monotoniadeltrabajo': 53,
        'psicosociales_sobrecargalaboral': 54,
        'psicosociales_minuciosidaddelatarea': 55,
        'psicosociales_altaresponsabilidad': 56,
        'psicosociales_autonomaenlatomadedecisiones': 57,
        'psicosociales_supervisinyestilosdedireccindeficiente': 58,
        'psicosociales_conflictoderol': 59,
        'psicosociales_faltadeclaridadenlasfunciones': 60,
        'psicosociales_incorrectadistribucindeltrabajo': 61,
        'psicosociales_turnosrotativos': 62,
        'psicosociales_relacinesinterpersonales': 63, 'psicosociales_relacionesinterpersonales': 63,
        'psicosociales_inestabilidadlaboral': 64,
        'psicosociales_amenazadelincuencial': 65,
    }
    
    center = Alignment(horizontal='center', vertical='center')
    for rkey, is_checked in riesgos.items():
        if not is_checked:
            continue
        parts = rkey.rsplit('_', 1)
        if len(parts) != 2 or not parts[1].startswith('h'):
            continue
        base_key = parts[0]
        hora = parts[1]
        row = riesgo_filas.get(base_key)
        col = h_col.get(hora)
        if row and col:
            ws3[f'{col}{row}'] = 'X'
            ws3[f'{col}{row}'].alignment = center
            ws3[f'{col}{row}'].font = Font(size=9)
    
    medidas_txt = g.get('medidas', g.get('medidas_preventivas', '')).replace('\n', ' - ')
    w(ws3, 'G67', medidas_txt)
    ws3.row_dimensions[67].height = 22
    ws3.row_dimensions[68].height = 22
    ws3.row_dimensions[69].height = 22
    ws3.row_dimensions[70].height = 22
    
    # ========================================
    # HOJA 4: EVALUACION 3/3
    # ========================================
    ws4 = wb.worksheets[3]
    
    empleos = h_sec.get('empleos', [])
    for idx, emp in enumerate(empleos[:20]):
        row = 7 + idx
        w(ws4, f'B{row}', emp.get('centro', emp.get('empresa', '')))
        w(ws4, f'J{row}', emp.get('cargo', emp.get('actividad', '')))
        if emp.get('estado') == 'ANTERIOR':
            ck(ws4, f'W{row}', True)
        elif emp.get('estado') == 'ACTUAL':
            ck(ws4, f'Y{row}', True)
        w(ws4, f'AA{row}', emp.get('tiempo', ''), 'center')
        inc = emp.get('inc_data', {})
        if inc.get('incidente'): ck(ws4, f'AC{row}', True)
        if inc.get('accidente'): ck(ws4, f'AE{row}', True)
        if inc.get('enfermedad'): ck(ws4, f'AH{row}', True)
        if inc.get('calif_si'): ck(ws4, f'AK{row}', True)
        if inc.get('calif_no'): ck(ws4, f'AM{row}', True)
        w(ws4, f'AO{row}', inc.get('fecha', ''), 'center')
        w(ws4, f'AR{row}', inc.get('especificar', ''))
        w(ws4, f'BE{row}', inc.get('observaciones', emp.get('observaciones', '')))
    
    extras = i_sec.get('actividades', [])
    for idx, act in enumerate(extras[:3]):
        row = 29 + idx
        desc = act.get('desc', act.get('descripcion', ''))
        tipo = act.get('tipo', '')
        texto = f'{tipo} {desc}'.strip() if tipo else desc
        w(ws4, f'B{row}', texto)
        w(ws4, f'BA{row}', act.get('fecha', ''), 'center')
    
    categorias = j.get('categorias', [])
    for idx, cat in enumerate(categorias[:6]):
        row = 35 + idx
        w(ws4, f'B{row}', cat.get('nombre', ''))
        w(ws4, f'M{row}', cat.get('fecha', ''), 'center')
        items = cat.get('items', [])
        res_str = ' - '.join(f"{it['nombre']}: {it['valor']} {it.get('unidad','')}" for it in items)
        w(ws4, f'T{row}', res_str, 'center')
    
    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.styles.colors import Color
    center = Alignment(horizontal='center', vertical='center')
    
    gray_dark = Color(rgb='FF808080')
    gray_light = Color(rgb='FFC0C0C0')
    thin_g = Side(style='thin', color=gray_light)
    thick_g = Side(style='thick', color=gray_dark)
    
    for row in range(45, 51):
        try:
            ws4.unmerge_cells(f'B{row}:P{row}')
        except:
            pass
        try:
            ws4.merge_cells(f'C{row}:P{row}')
        except:
            pass
        is_last = (row == 50)
        b_bottom = thick_g if is_last else thin_g
        ws4.cell(row=row, column=2).border = Border(left=thick_g, right=thin_g, top=thin_g, bottom=b_bottom)
        ws4.cell(row=row, column=3).border = Border(left=thin_g, top=thin_g, bottom=b_bottom)
        ws4.cell(row=row, column=16).border = Border(right=thin_g, top=thin_g, bottom=b_bottom)
        for col in range(4, 16):
            ws4.cell(row=row, column=col).border = Border(top=thin_g, bottom=b_bottom)
    
    diagnosticos = k.get('diagnosticos', [])
    for idx in range(6):
        row = 45 + idx
        ws4.cell(row=row, column=2).value = str(idx + 1)
        ws4.cell(row=row, column=2).alignment = center
        ws4.cell(row=row, column=2).font = Font(size=9)
        if idx < len(diagnosticos):
            dx = diagnosticos[idx]
            codigo = dx.get('codigo', '')
            ws4.cell(row=row, column=3).value = codigo
            ws4.cell(row=row, column=3).alignment = center
            ws4.cell(row=row, column=3).font = Font(size=9)
            desc = dx.get('descripcion', '').upper()
            w(ws4, f'Q{row}', desc, 'center')
            try:
                ws4[f'Q{row}'].font = Font(size=9)
            except:
                pass
            ck(ws4, f'BE{row}', dx.get('tipo') == 'PRE')
            ck(ws4, f'BJ{row}', dx.get('tipo') == 'DEF')
    
    from openpyxl.styles import Alignment
    center = Alignment(horizontal='center', vertical='center')
    
    ck(ws4, 'P53', aptitud == 'APTO')
    ws4['P53'].alignment = center
    ck(ws4, 'AE53', aptitud == 'APTO_OBS')
    ws4['AE53'].alignment = center
    ck(ws4, 'AS53', aptitud == 'APTO_LIM')
    ws4['AS53'].alignment = center
    ck(ws4, 'BB53', aptitud == 'NO_APTO')
    ws4['BB53'].alignment = center
    
    w(ws4, 'B54', l.get('observaciones', ''), 'center')
    ws4.row_dimensions[54].height = 25
    
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for row in [60, 61, 62, 63]:
        try:
            ws4.unmerge_cells(f'B{row}:BM{row}')
        except:
            pass
    try:
        ws4.merge_cells('B60:BM63')
    except:
        pass
    ws4.row_dimensions[60].height = 18
    ws4.row_dimensions[61].height = 18
    ws4.row_dimensions[62].height = 18
    ws4.row_dimensions[63].height = 18
    
    recs_list = m.get('estandar', [])
    recs_text = ' - '.join(recs_list)
    if m.get('medicas'):
        recs_text += ' - ' + m.get('medicas', '')
    if not recs_text:
        recs_text = (' - '.join(m.get('estandar',[])) + (' - ' + m.get('medicas','') if m.get('medicas') else '')) or m.get('descripcion','')
    
    ws4['B60'] = recs_text
    ws4['B60'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws4['B60'].font = Font(size=10)
    
    w(ws4, 'K73', prof_nombre, 'center')
    w(ws4, 'AD73', prof_codigo, 'center')

    # ========================================
    # INSERTAR FIRMA EN HOJA 4 (sección N/O)
    # ========================================
    if firma_b64 and (modo in ('formulario', 'todo')):
        insertar_firma(ws4, firma_b64, col_letra='AM', fila=73, ancho_px=155, alto_px=55)
        print('✅ Firma insertada en FORMULARIO hoja 4')

    # ========================================
    # ELIMINAR HOJAS SEGÚN MODO
    # ========================================
    if modo == 'certificado':
        for sheet_name in list(wb.sheetnames[1:]):
            del wb[sheet_name]
    elif modo == 'formulario':
        del wb[wb.sheetnames[0]]
    
    # GUARDAR
    xlsx_path = output_path.replace('.pdf', '.xlsx')
    wb.save(xlsx_path)
    print(f"✅ Excel guardado: {xlsx_path}")
    
    if output_path.endswith('.pdf'):
        outdir = os.path.dirname(output_path) or '.'
        cmd = ['soffice', '--headless', '--calc', '--convert-to', 'pdf',
               xlsx_path, '--outdir', outdir]
        env = dict(os.environ, HOME='/tmp')
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, env=env)
        if result.returncode == 0:
            generated = xlsx_path.replace('.xlsx', '.pdf')
            if generated != output_path:
                os.rename(generated, output_path)
            print(f"✅ PDF generado: {output_path}")
        else:
            print(f"❌ Error PDF: {result.stderr}")
    
    return output_path
