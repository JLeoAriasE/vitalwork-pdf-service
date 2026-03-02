#!/usr/bin/env python3
"""
fill_formulario.py - Llena el formulario oficial MSP con datos del paciente
Recibe JSON con datos, llena el Excel, exporta PDF
"""
from openpyxl import load_workbook
from openpyxl.styles.borders import Border
import json, sys, os, subprocess, tempfile

def fill_formulario(data, template_path, output_path, modo='todo'):
    """
    data: dict con todos los datos de la ficha ocupacional
    template_path: ruta al formulario_final.xlsx (plantilla sin diagonales)
    output_path: ruta de salida (.pdf o .xlsx)
    modo: 'certificado' (solo hoja 1), 'formulario' (hojas 2,3,4), 'todo' (todas)
    """
    wb = load_workbook(template_path)
    
    # Helpers
    def w(ws, cell, val, align='left', font_size=9):
        """Escribir valor en celda, maneja celdas fusionadas, aplica wrap_text"""
        if val is None or val == '':
            return
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Alignment, Font
        wrap = Alignment(horizontal=align, vertical='center', wrap_text=True)
        fnt = Font(size=font_size)
        target = ws[cell]
        if isinstance(target, MergedCell):
            for m in ws.merged_cells.ranges:
                if cell in m:
                    tl = ws.cell(row=m.min_row, column=m.min_col)
                    tl.value = str(val)
                    tl.alignment = wrap
                    tl.font = fnt
                    return
        else:
            ws[cell] = str(val)
            target = ws[cell]
            target.alignment = wrap
            target.font = fnt
    
    def ck(ws, cell, condition):
        """Checkbox: X si true, vacío si false"""
        from openpyxl.cell.cell import MergedCell
        target = ws[cell]
        val = 'X' if condition else ''
        if isinstance(target, MergedCell):
            for m in ws.merged_cells.ranges:
                if cell in m:
                    ws.cell(row=m.min_row, column=m.min_col).value = val
                    return
        else:
            ws[cell] = val
    
    # ========================================
    # Datos comunes
    # ========================================
    # La app envía: { a:{...}, b:{...}, ..., n:{...}, paciente:{cedula,fecha_nacimiento}, empresa:{nombre,ruc,ciiu} }
    # O la estructura legacy: { paciente:{}, empresa:{}, ficha:{a,b,...}, profesional:{} }
    # Soportar ambas
    # ========================================
    
    if 'ficha' in data:
        # Formato legacy
        pac = data.get('paciente', {})
        ec = data.get('empresa', {})
        fo = data.get('ficha', {})
        prof_nombre = data.get('profesional', {}).get('nombre', '')
        prof_codigo = data.get('profesional', {}).get('codigo', '')
    else:
        # Formato directo desde FO.data de la app
        fo = data
        pac = data.get('_paciente', {})
        ec = data.get('_empresa', {})
        prof_nombre = ''
        prof_codigo = ''
    
    a = fo.get('a', {})
    b = fo.get('b', {})
    c = fo.get('c', {})
    d = fo.get('d', {})
    e = fo.get('e', {})
    f = fo.get('f', {})
    g = fo.get('g', {})
    h_sec = fo.get('h', {})
    i_sec = fo.get('i', {})
    j = fo.get('j', {})
    k = fo.get('k', {})
    l = fo.get('l', {})
    m = fo.get('m', {})
    n = fo.get('n', {})
    r = fo.get('r', {})
    
    # Extraer datos del paciente (de sección A o de paciente)
    ap1 = a.get('ap1', pac.get('apellido1', ''))
    ap2 = a.get('ap2', pac.get('apellido2', ''))
    n1 = a.get('n1', pac.get('nombre1', ''))
    n2 = a.get('n2', pac.get('nombre2', ''))
    cedula = a.get('nform', pac.get('cedula', ''))
    sexo_raw = a.get('sexo', pac.get('sexo', ''))
    sexo = 'Masculino' if sexo_raw in ('M', 'Masculino') else 'Femenino' if sexo_raw in ('F', 'Femenino') else sexo_raw
    sexo_corto = 'M' if sexo == 'Masculino' else 'F' if sexo == 'Femenino' else sexo_raw
    fecha_nac = pac.get('fecha_nacimiento', '')
    edad = pac.get('edad', '')
    
    # Empresa
    ruc = a.get('ruc', ec.get('ruc', ''))
    ciiu = a.get('ciiu', ec.get('ciiu', ''))
    empresa_nombre = a.get('empresa', ec.get('nombre', ''))
    centro_trabajo = a.get('centro', empresa_nombre)
    
    # Profesional (de sección N o legacy)
    if not prof_nombre:
        prof_nombre = n.get('medico', '')
    if not prof_codigo:
        prof_codigo = n.get('codigo', '')
    
    # Tipo evaluación y aptitud
    tipo_eval_raw = b.get('tipo_eval', b.get('tipo', ''))
    tipo_map = {'INGRESO':'INGRESO','PERIÓDICO':'PERIODICO','PERIODICO':'PERIODICO','REINTEGRO':'REINTEGRO','RETIRO':'RETIRO'}
    tipo_eval = tipo_map.get(tipo_eval_raw.upper(), tipo_eval_raw.upper()) if tipo_eval_raw else ''
    
    aptitud_raw = l.get('aptitud', '')
    aptitud_map = {'APTO':'APTO','APTO EN OBSERVACIÓN':'APTO_OBS','APTO CON LIMITACIONES':'APTO_LIM','NO APTO':'NO_APTO','APTO_OBS':'APTO_OBS','APTO_LIM':'APTO_LIM','NO_APTO':'NO_APTO'}
    aptitud = aptitud_map.get(aptitud_raw, aptitud_raw)
    
    # Parse fecha
    def parse_fecha(f):
        if not f: return ('','','')
        parts = f.split('-') if '-' in f else f.split('/')
        if len(parts) == 3:
            return (parts[0], parts[1], parts[2])
        return ('','','')
    
    fecha_eval = b.get('fecha_atencion', b.get('fecha', ''))
    f_y, f_m, f_d = parse_fecha(fecha_eval)
    fn_y, fn_m, fn_d = parse_fecha(fecha_nac)
    
    # ========================================
    # HOJA 1: CERTIFICADO
    # ========================================
    ws1 = wb['CERTIFICADO']
    
    # Sección A - Datos establecimiento
    # INSTITUCIÓN DEL SISTEMA → vacío (es para sistema público)
    # ESTABLECIMIENTO/CENTRO DE TRABAJO → empresa del colaborador
    w(ws1, 'L4', ruc, 'center')
    w(ws1, 'R4', ciiu, 'center')
    w(ws1, 'V4', empresa_nombre, 'center')  # Empresa va en ESTABLECIMIENTO
    w(ws1, 'AC4', a.get('nform', cedula), 'center')  # N° Formulario
    w(ws1, 'AI4', a.get('narch', cedula), 'center')  # N° Archivo
    
    # Nombres centrados
    w(ws1, 'A6', ap1, 'center')
    w(ws1, 'J6', ap2, 'center')
    w(ws1, 'Q6', n1, 'center')
    w(ws1, 'X6', n2, 'center')
    w(ws1, 'AD6', sexo_corto, 'center')
    w(ws1, 'AG6', b.get('puesto_ciuo', ''), 'center')
    ws1.row_dimensions[6].height = 25  # Puesto wrap
    
    # Sección B - Fecha de emisión (cada dígito en su cuadrito, centrado)
    from openpyxl.styles import Alignment
    center = Alignment(horizontal='center', vertical='center')
    
    # K10:L10=año, M10=mes1, N10=mes2, O10=dia1, P10=dia2
    w(ws1, 'K10', f_y, 'center')
    ws1['K10'].alignment = center
    w(ws1, 'M10', f_m[0] if len(f_m) >= 1 else '')
    ws1['M10'].alignment = center
    w(ws1, 'N10', f_m[1] if len(f_m) >= 2 else '')
    ws1['N10'].alignment = center
    w(ws1, 'O10', f_d[0] if len(f_d) >= 1 else '')
    ws1['O10'].alignment = center
    w(ws1, 'P10', f_d[1] if len(f_d) >= 2 else '')
    ws1['P10'].alignment = center
    
    # Tipo de evaluación - X centrada en checkbox
    ck(ws1, 'L12', tipo_eval == 'INGRESO')
    ws1['L12'].alignment = center
    ck(ws1, 'U12', tipo_eval == 'PERIODICO')
    ws1['U12'].alignment = center
    ck(ws1, 'AC12', tipo_eval == 'REINTEGRO')
    ws1['AC12'].alignment = center
    ck(ws1, 'AI12', tipo_eval == 'RETIRO')
    ws1['AI12'].alignment = center
    
    # Sección C - Aptitud (X centrada en checkbox correcto)
    # I17=checkbox APTO, S17:T17=checkbox APTO_OBS, AC17:AD17=checkbox APTO_LIM, AK17:AL17=checkbox NO_APTO
    from openpyxl.styles import Alignment
    center = Alignment(horizontal='center', vertical='center')
    
    ck(ws1, 'I17', aptitud == 'APTO')
    ws1['I17'].alignment = center
    ck(ws1, 'S17', aptitud == 'APTO_OBS')
    ws1['S17'].alignment = center
    ck(ws1, 'AC17', aptitud == 'APTO_LIM')
    ws1['AC17'].alignment = center
    ck(ws1, 'AK17', aptitud == 'NO_APTO')
    ws1['AK17'].alignment = center
    
    # Observaciones aptitud - va en fila 19 (debajo de "DETALLE DE OBSERVACIONES:" fila 18)
    obs_text = l.get('observaciones', '')
    w(ws1, 'A19', obs_text, 'center')
    ws1.row_dimensions[19].height = 25
    ws1.row_dimensions[20].height = 25
    
    # Sección D - Recomendaciones (Certificado)
    # Fusionar filas 25-27 (3 filas) para tener más espacio
    from openpyxl.styles import Alignment, Font
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Desmerge filas individuales y fusionar en bloque
    for row in [25, 26, 27]:
        try:
            ws1.unmerge_cells(f'A{row}:AL{row}')
        except:
            pass
    try:
        ws1.merge_cells('A25:AL27')
    except:
        pass
    
    # Aumentar altura de filas para más espacio
    ws1.row_dimensions[25].height = 22
    ws1.row_dimensions[26].height = 22
    ws1.row_dimensions[27].height = 22
    
    # Unir recomendaciones separadas por guión
    recs_list = m.get('estandar', [])
    recs_text = ' - '.join(recs_list)
    if m.get('medicas'):
        recs_text += ' - ' + m.get('medicas', '')
    if not recs_text:
        recs_text = (' - '.join(m.get('estandar',[])) + (' - ' + m.get('medicas','') if m.get('medicas') else '')) or m.get('descripcion','')
    
    ws1['A25'] = recs_text
    ws1['A25'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1['A25'].font = Font(size=10)
    
    # Sección E - Profesional (datos en celdas vacías, no en labels)
    # Labels: A33:D33, L33:N33, T33:V33
    # Datos: E33:K33=nombre, O33:S33=código
    w(ws1, 'E33', prof_nombre, 'center')   # cols 4-10
    w(ws1, 'O33', prof_codigo, 'center')   # cols 14-18
    
    # ========================================
    # HOJA 2: EVALUACION 1/3
    # ========================================
    ws2 = wb.worksheets[1]
    from openpyxl.styles import Alignment
    center = Alignment(horizontal='center', vertical='center')
    
    # Sección A - Datos establecimiento
    # Fila 4 (xlrd) = Fila 5 (openpyxl) = datos (debajo de labels fila 4)
    # INSTITUCIÓN → vacío, ESTABLECIMIENTO → empresa
    w(ws2, 'Q5', ruc, 'center')          # cols 16-24 = RUC
    w(ws2, 'Z5', ciiu, 'center')         # cols 25-27 = CIIU
    w(ws2, 'AC5', empresa_nombre, 'center')  # cols 28-40 = ESTABLECIMIENTO
    w(ws2, 'AP5', a.get('nform', cedula), 'center')  # N° Historia Clínica
    w(ws2, 'BB5', a.get('narch', cedula), 'center')  # N° Archivo
    
    # Nombres - fila 7 (xlrd) = fila 8 (openpyxl) = datos
    # Labels en filas 6-7 (openpyxl), datos en fila 8
    w(ws2, 'B8', ap1, 'center')          # cols 1-17
    w(ws2, 'S8', ap2, 'center')          # cols 18-32
    w(ws2, 'AH8', n1, 'center')          # cols 33-42
    w(ws2, 'AR8', n2, 'center')          # cols 43-58
    
    # Fila 12 (xlrd) = fila 13 (openpyxl) = datos checkboxes
    # Atención prioritaria checkboxes
    # cols 1-5=Embarazada, 6-9=Discapacidad, 10-12=Catastrófica, 13-16=Lactancia, 17-19=Adulto Mayor
    
    # Sexo - checkboxes en fila 13
    # cols 20-22=hombre, 23-24=mujer (labels en fila 11, checkboxes en fila 13)
    ck(ws2, 'U13', sexo == 'Masculino')    # col 20
    ws2['U13'].alignment = center
    ck(ws2, 'X13', sexo == 'Femenino')     # col 23
    ws2['X13'].alignment = center
    
    # Fecha nacimiento - fila 13
    # cols 25-27=año, 28-29=mes, 30=día
    w(ws2, 'Z13', fn_y, 'center')       # col 25 = año
    ws2['Z13'].alignment = center
    w(ws2, 'AC13', fn_m, 'center')      # col 28 = mes
    ws2['AC13'].alignment = center
    w(ws2, 'AE13', fn_d, 'center')      # col 30 = día
    ws2['AE13'].alignment = center
    
    # Edad - fila 13, cols 31-32
    w(ws2, 'AF13', edad, 'center')      # col 31
    ws2['AF13'].alignment = center
    
    # Grupo sanguíneo - fila 13, cols 33-40
    w(ws2, 'AH13', a.get('grupo_sang', ''), 'center')
    
    # Lateralidad - fila 13, cols 41-58
    lat_raw = a.get('lateralidad', '')
    lat_text = 'Derecha' if lat_raw == 'Diestro' else 'Izquierda' if lat_raw == 'Zurdo' else lat_raw
    w(ws2, 'AP13', lat_text, 'center')
    
    # Sección B - Motivo consulta
    # Puesto de trabajo - fila 16 (openpyxl), cols 9-23 (dato)
    w(ws2, 'J16', b.get('puesto_ciuo', ''), 'center')
    
    # Fecha de atención - fila 16, cols 33-58
    w(ws2, 'AH16', b.get('fecha_atencion', ''), 'center')
    
    # Fecha de ingreso - fila 17 (openpyxl 17), cols 1-28
    fi_fecha = b.get('fecha_ingreso', '')
    w(ws2, 'B17', fi_fecha, 'center')
    
    # Fecha de reintegro - fila 17, cols 29-37
    w(ws2, 'AD17', b.get('fecha_reintegro', ''), 'center')
    
    # Fecha último día laboral - fila 17, cols 38-58
    w(ws2, 'AM17', b.get('fecha_ultimo_dia', ''), 'center')
    
    # Tipo evaluación - fila 19 (xlrd 18), checkboxes en merges vacíos
    # cols 6-13=INGRESO, 25-28=PERIÓDICO, 34-37=REINTEGRO, 52-58=RETIRO
    ck(ws2, 'G19', tipo_eval == 'INGRESO')     # col 6
    ws2['G19'].alignment = center
    ck(ws2, 'Z19', tipo_eval == 'PERIODICO')    # col 25
    ws2['Z19'].alignment = center
    ck(ws2, 'AI19', tipo_eval == 'REINTEGRO')   # col 34
    ws2['AI19'].alignment = center
    ck(ws2, 'BA19', tipo_eval == 'RETIRO')      # col 52
    ws2['BA19'].alignment = center
    
    w(ws2, 'B21', b.get('observacion', ''), 'center')
    
    # Sección C - Antecedentes
    # Antecedentes: reemplazar saltos de línea por separador visible
    ant_clin = c.get('ant_clinicos', c.get('clinicos', '')).replace('\n', ' - ')
    ant_fam = c.get('ant_familiares', c.get('familiares', '')).replace('\n', ' - ')
    w(ws2, 'B24', ant_clin)
    ws2.row_dimensions[24].height = 35
    w(ws2, 'B26', ant_fam)
    ws2.row_dimensions[26].height = 35
    
    # Sección D - Enfermedad actual
    w(ws2, 'B51', d.get('descripcion', ''))
    ws2.row_dimensions[51].height = 35
    
    # Sección E - Constantes vitales
    w(ws2, 'B55', e.get('temp', e.get('temperatura', '')), 'center')
    pa_val = str(e.get('pa_s',''))+('/'+str(e.get('pa_d','')) if e.get('pa_d') else '') if e.get('pa_s') else e.get('presion','')
    w(ws2, 'I55', pa_val, 'center')
    w(ws2, 'P55', e.get('fc', ''), 'center')
    w(ws2, 'W55', e.get('fr', ''), 'center')
    w(ws2, 'AC55', e.get('sat', e.get('sat_o2', '')), 'center')
    w(ws2, 'AG55', e.get('peso', ''), 'center')
    w(ws2, 'AI55', e.get('talla', ''), 'center')
    w(ws2, 'AM55', e.get('imc', ''), 'center')
    w(ws2, 'AS55', e.get('perim_abd', e.get('perimetro', '')), 'center')
    
    # Sección F - Examen físico
    # La app guarda: f.hallazgos = {'1a':true, '10c':true, '12b':true}
    # Clave = número_región + letra_item
    # Mapeo hallazgo → celda checkbox en hoja 2 (celdas vacías de merge)
    hallazgos = f.get('hallazgos', {})
    hall_cell_map = {
        # Región 1: Piel (cols 7-8)
        '1a': 'H59', '1b': 'H60', '1c': 'H61',
        # Región 2: Ojos
        '2a': 'H63', '2b': 'H64', '2c': 'H65', '2d': 'H66', '2e': 'H68',
        # Región 3: Oído (cols 19-20)
        '3a': 'T59', '3b': 'T60', '3c': 'T62',
        # Región 4: Orofaringe
        '4a': 'T63', '4b': 'T64', '4c': 'T65', '4d': 'T66', '4e': 'T68',
        # Región 5: Nariz (cols 28)
        '5a': 'AC59', '5b': 'AC60', '5c': 'AC62', '5d': 'AC63',
        # Región 6: Cuello
        '6a': 'AC64', '6b': 'AC65',
        # Región 7: Tórax
        '7a': 'AC66', '7b': 'AC69',
        # Región 8: Tórax Órganos (cols 35)
        '8a': 'AJ59', '8b': 'AJ60', '8c': 'AJ61',
        # Región 9: Abdomen
        '9a': 'AJ62', '9b': 'AJ63',
        # Región 10: Columna
        '10a': 'AJ64', '10b': 'AJ66', '10c': 'AJ68',
        # Región 11: Pelvis (cols 57-58)
        '11a': 'BF59', '11b': 'BF60',
        # Región 12: Extremidades
        '12a': 'BF62', '12b': 'BF63', '12c': 'BF64',
        # Región 13: Neurológico
        '13a': 'BF65', '13b': 'BF66', '13c': 'BF67', '13d': 'BF68',
    }
    for key, cell in hall_cell_map.items():
        if hallazgos.get(key):
            ck(ws2, cell, True)
    
    w(ws2, 'B74', f.get('observaciones', f.get('observacion', '')))
    
    # ========================================
    # HOJA 3: FACTORES DE RIESGO (2/3)
    # ========================================
    ws3 = wb.worksheets[2]
    
    w(ws3, 'G2', b.get('puesto_ciuo', ''), 'center', 10)
    ws3.row_dimensions[2].height = 25
    
    # Riesgos: la app guarda {'fisicos_ruido_h1':true, 'quimicos_polvos_h2':true, ...}
    # Formato: categoria_itemnormalizado_h{hora}
    riesgos = g.get('riesgos', {})
    
    # Columnas horas: openpyxl G=h1, I=h2, K=h3, M=h4, N=h5, O=h6, P=h7
    h_col = {'h1':'G','h2':'I','h3':'K','h4':'M','h5':'N','h6':'O','h7':'P'}
    
    # Mapeo: nombre normalizado en la app → fila openpyxl
    riesgo_filas = {
        # FÍSICO
        'fisicos_temperaturasaltas': 6, 'fisicos_temperaturasbajas': 7,
        'fisicos_radiacinionizante': 8, 'fisicos_radiacinnoionizante': 9,
        'fisicos_ruido': 10, 'fisicos_vibracin': 11,
        'fisicos_iluminacin': 12, 'fisicos_ventilacin': 13,
        'fisicos_fluidoelctrico': 14,
        # SEGURIDAD - LOCATIVOS
        'seguridad_faltadesealizacinaseodesorden': 16,
        # SEGURIDAD - MECÁNICOS
        'seguridad_atrapamientoentremquinasyosuperficies': 18,
        'seguridad_atrapamientoentreojetos': 18,
        'seguridad_atrapamientoentreobetos': 18,
        'seguridad_cadadeojetos': 19, 'seguridad_cadadeobetos': 19,
        'seguridad_cadasalmismonivel': 20,
        'seguridad_cadasadiferentenivel': 21,
        'seguridad_pinchazos': 22, 'seguridad_cortes': 23,
        'seguridad_choquescolisinvehicular': 24,
        'seguridad_atropellamientosporvehculos': 25,
        'seguridad_proyeccindefluidos': 26,
        'seguridad_proyeccindepartculasfragmentos': 27,
        'seguridad_contactoconsuperficiesdetrabajo': 28,
        # ELÉCTRICOS
        'seguridad_contactoelctrico': 29,
        # QUÍMICO
        'quimicos_polvos': 31, 'quimicos_slidos': 32, 'quimicos_solidos': 32,
        'quimicos_humos': 33, 'quimicos_lquidos': 34, 'quimicos_liquidos': 34,
        'quimicos_vapores': 35, 'quimicos_aerosoles': 36,
        'quimicos_neblinas': 37, 'quimicos_gaseosos': 38,
        # BIOLÓGICO
        'biologicos_virus': 40, 'biologicos_hongos': 41,
        'biologicos_bacterias': 42, 'biologicos_parsitos': 43, 'biologicos_parasitos': 43,
        'biologicos_exposicinavectores': 44, 'biologicos_exposicionavectores': 44,
        'biologicos_exposicinaanimalesselvaticos': 45, 'biologicos_exposicionaanimalesselvaticos': 45,
        # ERGONÓMICO
        'ergonomicos_manejomanualdecargas': 47,
        'ergonomicos_movimientosrepetitivos': 48, 'ergonomicos_movimientorepetitivos': 48,
        'ergonomicos_postursforzadas': 49, 'ergonomicos_posturasforzadas': 49,
        'ergonomicos_trabajosconpvd': 50,
        'ergonomicos_diseoinadecuadodelpuesto': 51, 'ergonomicos_disenoinadecuadodelpuesto': 51,
        # PSICOSOCIAL
        'psicosociales_monotoniadeltrabajo': 53,
        'psicosociales_sobrecargalaboral': 54,
        'psicosociales_minuciosidaddelatarea': 55,
        'psicosociales_altaresponsabilidad': 56,
        'psicosociales_autonomaenlatomadedecisiones': 57,
        'psicosociales_supervisinyestilosdedireccindeficiente': 58,
        'psicosociales_conflictoderol': 59,
        'psicosociales_faltadeclaridadenlasfunciones': 60,
        'psicosociales_incorrectadistribucindeltrabajo': 61,
        'psicosociales_turnosrotativos': 62,
        'psicosociales_relacinesinterpersonales': 63, 'psicosociales_relacionesinterpersonales': 63,
        'psicosociales_inestabilidadlaboral': 64,
        'psicosociales_amenazadelincuencial': 65,
    }
    
    center = Alignment(horizontal='center', vertical='center')
    for rkey, is_checked in riesgos.items():
        if not is_checked:
            continue
        # Separar: 'fisicos_ruido_h1' → base='fisicos_ruido', hora='h1'
        parts = rkey.rsplit('_', 1)
        if len(parts) != 2 or not parts[1].startswith('h'):
            continue
        base_key = parts[0]
        hora = parts[1]  # 'h1', 'h2', etc.
        row = riesgo_filas.get(base_key)
        col = h_col.get(hora)
        if row and col:
            ws3[f'{col}{row}'] = 'X'
            ws3[f'{col}{row}'].alignment = center
            ws3[f'{col}{row}'].font = Font(size=9)
    
    # Medidas preventivas
    medidas_txt = g.get('medidas', g.get('medidas_preventivas', '')).replace('\n', ' - ')
    w(ws3, 'A68', medidas_txt)
    
    # ========================================
    # HOJA 4: EVALUACION 3/3
    # ========================================
    ws4 = wb.worksheets[3]
    
    # Sección H - Historial laboral
    empleos = h_sec.get('empleos', [])
    for idx, emp in enumerate(empleos[:20]):  # max 20 filas
        row = 6 + idx
        w(ws4, f'B{row}', emp.get('centro', emp.get('empresa', '')))
        w(ws4, f'J{row}', emp.get('cargo', emp.get('actividad', '')))
    
    # Sección I - Actividades extra
    extras = i_sec.get('actividades', [])
    for idx, act in enumerate(extras[:3]):
        row = 29 + idx
        w(ws4, f'B{row}', act.get('descripcion', ''))
    
    # Sección J - Resultados exámenes
    categorias = j.get('categorias', [])
    for idx, cat in enumerate(categorias[:6]):
        row = 35 + idx
        w(ws4, f'B{row}', cat.get('nombre', ''))
        w(ws4, f'M{row}', cat.get('fecha', ''), 'center')
        # Resultados como string
        items = cat.get('items', [])
        res_str = ' - '.join(f"{it['nombre']}: {it['valor']} {it.get('unidad','')}" for it in items)
        w(ws4, f'T{row}', res_str, 'center')
    
    # Sección K - Diagnósticos
    # Separar B (número fila) y C:P (código CIE-10)
    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.styles.colors import Color
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    gray_dark = Color(rgb='FF808080')
    gray_light = Color(rgb='FFC0C0C0')
    thin_g = Side(style='thin', color=gray_light)
    thick_g = Side(style='thick', color=gray_dark)
    
    for row in range(45, 51):
        # Desmerge B:P original
        try:
            ws4.unmerge_cells(f'B{row}:P{row}')
        except:
            pass
        # Nuevo merge C:P para CIE-10
        try:
            ws4.merge_cells(f'C{row}:P{row}')
        except:
            pass
        
        is_last = (row == 50)
        b_bottom = thick_g if is_last else thin_g
        
        # B (número): bordes
        ws4.cell(row=row, column=2).border = Border(
            left=thick_g, right=thin_g, top=thin_g, bottom=b_bottom)
        
        # C (inicio CIE)
        ws4.cell(row=row, column=3).border = Border(
            left=thin_g, top=thin_g, bottom=b_bottom)
        
        # P (fin CIE, col 16): borde derecho
        ws4.cell(row=row, column=16).border = Border(
            right=thin_g, top=thin_g, bottom=b_bottom)
        
        # Bordes intermedios D-O (cols 4-15)
        for col in range(4, 16):
            ws4.cell(row=row, column=col).border = Border(
                top=thin_g, bottom=b_bottom)
    
    diagnosticos = k.get('diagnosticos', [])
    for idx in range(6):
        row = 45 + idx
        # Número de fila siempre
        ws4.cell(row=row, column=2).value = str(idx + 1)
        ws4.cell(row=row, column=2).alignment = center
        ws4.cell(row=row, column=2).font = Font(size=9)
        
        if idx < len(diagnosticos):
            dx = diagnosticos[idx]
            codigo = dx.get('codigo', '')
            # CIE-10 en C:P
            ws4.cell(row=row, column=3).value = codigo
            ws4.cell(row=row, column=3).alignment = center
            ws4.cell(row=row, column=3).font = Font(size=9)
            # Descripción
            desc = dx.get('descripcion', '').upper()
            w(ws4, f'Q{row}', desc, 'center')
            try:
                ws4[f'Q{row}'].font = Font(size=9)
            except:
                pass
            ck(ws4, f'BE{row}', dx.get('tipo') == 'PRE')
            ck(ws4, f'BJ{row}', dx.get('tipo') == 'DEF')
    
    # Sección L - Aptitud
    # xlrd fila 52: cols 15-16=APTO cb, 30-31=APTO_OBS cb, 44-45=APTO_LIM cb, 53+=NO_APTO cb
    # openpyxl: P53, AE53, AS53, BB53
    from openpyxl.styles import Alignment
    center = Alignment(horizontal='center', vertical='center')
    
    ck(ws4, 'P53', aptitud == 'APTO')
    ws4['P53'].alignment = center
    ck(ws4, 'AE53', aptitud == 'APTO_OBS')
    ws4['AE53'].alignment = center
    ck(ws4, 'AS53', aptitud == 'APTO_LIM')
    ws4['AS53'].alignment = center
    ck(ws4, 'BB53', aptitud == 'NO_APTO')
    ws4['BB53'].alignment = center
    
    w(ws4, 'B54', l.get('observaciones', ''), 'center')
    ws4.row_dimensions[54].height = 25  # Dentro de sección L, debajo de APTO/NO APTO
    
    # Sección M - Recomendaciones (Hoja 4)
    # Fusionar filas 60-63 (4 filas disponibles)
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    for row in [60, 61, 62, 63]:
        try:
            ws4.unmerge_cells(f'B{row}:BM{row}')
        except:
            pass
    try:
        ws4.merge_cells('B60:BM63')
    except:
        pass
    
    # Aumentar altura
    ws4.row_dimensions[60].height = 18
    ws4.row_dimensions[61].height = 18
    ws4.row_dimensions[62].height = 18
    ws4.row_dimensions[63].height = 18
    
    recs_list = m.get('estandar', [])
    recs_text = ' - '.join(recs_list)
    if m.get('medicas'):
        recs_text += ' - ' + m.get('medicas', '')
    if not recs_text:
        recs_text = (' - '.join(m.get('estandar',[])) + (' - ' + m.get('medicas','') if m.get('medicas') else '')) or m.get('descripcion','')
    
    ws4['B60'] = recs_text
    ws4['B60'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws4['B60'].font = Font(size=10)
    
    # Sección O - Profesional
    # Labels en cols 1-9, 20-28, 33-38 (merge filas 73-74)
    # Datos en cols 10-19, 29-32 (merge filas 73-74)
    # openpyxl: K73=nombre, AD73=código
    w(ws4, 'K73', prof_nombre, 'center')   # cols 10-19
    w(ws4, 'AD73', prof_codigo, 'center')  # cols 29-32
    
    # ========================================
    # ELIMINAR HOJAS SEGÚN MODO
    # ========================================
    if modo == 'certificado':
        # Solo hoja 1 (CERTIFICADO) - eliminar hojas 2,3,4
        for sheet_name in list(wb.sheetnames[1:]):
            del wb[sheet_name]
    elif modo == 'formulario':
        # Solo hojas 2,3,4 - eliminar hoja 1 (CERTIFICADO)
        del wb[wb.sheetnames[0]]
    # modo 'todo' = no eliminar nada
    
    # GUARDAR
    # ========================================
    xlsx_path = output_path.replace('.pdf', '.xlsx')
    wb.save(xlsx_path)
    print(f"✅ Excel guardado: {xlsx_path}")
    
    # Si piden PDF, convertir con LibreOffice
    if output_path.endswith('.pdf'):
        outdir = os.path.dirname(output_path) or '.'
        cmd = ['soffice', '--headless', '--calc', '--convert-to', 'pdf',
               xlsx_path, '--outdir', outdir]
        env = dict(os.environ, HOME='/tmp')
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, env=env)
        if result.returncode == 0:
            # Rename to expected output
            generated = xlsx_path.replace('.xlsx', '.pdf')
            if generated != output_path:
                os.rename(generated, output_path)
            print(f"✅ PDF generado: {output_path}")
        else:
            print(f"❌ Error PDF: {result.stderr}")
    
    return output_path


# === TEST ===
if __name__ == '__main__':
    # Test con formato FO.data directo (como envía la app)
    test_data = {
        '_paciente': {
            'cedula': '0705191229',
            'fecha_nacimiento': '1990-05-15',
            'edad': '35',
        },
        '_empresa': {
            'nombre': 'CONSTRUCTORA MACHALA S.A.',
            'ruc': '0791234567001',
            'ciiu': 'F4100',
        },
        'a': {
            'empresa': 'CONSTRUCTORA MACHALA S.A.',
            'ruc': '0791234567001',
            'ciiu': 'F4100',
            'centro': 'CONSTRUCTORA MACHALA S.A.',
            'nform': '0705191229',
            'narch': '0705191229',
            'ap1': 'ARIAS',
            'ap2': 'ESPINOZA',
            'n1': 'JORGE',
            'n2': 'LEONARDO',
            'sexo': 'M',
            'cargo': '(8121) OPERADOR DE MAQUINARIA MINERA',
            'grupo_sang': 'O+',
            'lateralidad': 'Diestro',
            'atencion_prioritaria': [],
        },
        'b': {
            'puesto_ciuo': '(8121) OPERADOR DE MAQUINARIA MINERA',
            'fecha_atencion': '2026-03-01',
            'tipo_eval': 'INGRESO',
            'fecha_ingreso': '2026-03-01',
            'fecha_reintegro': '',
            'fecha_ultimo_dia': '',
            'observacion': 'EVALUACIÓN PRE-OCUPACIONAL PARA INGRESO A OBRA DE CONSTRUCCIÓN VIAL',
        },
        'c': {
            'ant_clinicos': 'APP: NIEGA\nAQx: APENDICECTOMÍA 2018\nALERGIAS: PENICILINA\nAPsi: NIEGA\nTRAUMATOLÓGICOS: FRACTURA RADIO IZQUIERDO 2015, RESUELTA SIN SECUELAS',
            'ant_familiares': 'PADRE: HTA, DM2\nMADRE: HIPOTIROIDISMO\nHERMANOS: SANOS',
            'aut_transfusion': 'Sí',
            'trat_hormonal': 'No',
            'sustancias': [
                {'nombre': 'Tabaco', 'estado': 'Pasivo', 'consumo': '36', 'abstinencia': '24'},
                {'nombre': 'Alcohol', 'estado': 'Activo', 'consumo': '12', 'abstinencia': ''},
                {'nombre': 'Otras drogas', 'estado': 'Nunca', 'consumo': '', 'abstinencia': ''},
            ],
            'medicacion': 'LOSARTÁN 50MG QD - METFORMINA 850MG BID',
        },
        'd': {
            'descripcion': 'PACIENTE MASCULINO DE 35 AÑOS ACUDE PARA EVALUACIÓN PREOCUPACIONAL DE INGRESO. REFIERE ENCONTRARSE EN BUEN ESTADO DE SALUD GENERAL. NIEGA SINTOMATOLOGÍA ACTUAL. REFIERE DOLOR LUMBAR OCASIONAL CON ESFUERZOS FÍSICOS QUE CEDE CON REPOSO.',
        },
        'e': {
            'temp': '36.5',
            'pa_s': '130',
            'pa_d': '85',
            'fc': '78',
            'fr': '18',
            'sat': '97',
            'peso': '82',
            'talla': '175',
            'imc': '26.8',
            'clasif_imc': 'Sobrepeso',
            'perim_abd': '92',
        },
        'f': {
            'hallazgos': {
                '1a': True,   # Cicatrices (piel)
                '10c': True,  # Dolor (columna)
                '12b': True,  # M. Superiores
            },
            'observaciones': '1A CICATRIZ QUIRÚRGICA EN FOSA ILÍACA DERECHA DE APROX 8CM POR APENDICECTOMÍA. 10C DOLOR A LA PALPACIÓN EN REGIÓN LUMBAR L4-L5 SIN IRRADIACIÓN. 12B LIMITACIÓN LEVE EN PRONACIÓN DE MUÑECA IZQUIERDA POR ANTECEDENTE DE FRACTURA.',
        },
        'g': {
            'riesgos': {
                'fisicos_ruido_h1': True, 'fisicos_ruido_h2': True, 'fisicos_ruido_h3': True, 'fisicos_ruido_h4': True,
                'fisicos_vibracion_h1': True, 'fisicos_vibracion_h2': True,
                'ergonomicos_manejomanualdecargas_h1': True, 'ergonomicos_manejomanualdecargas_h2': True, 'ergonomicos_manejomanualdecargas_h3': True,
                'ergonomicos_postursforzadas_h1': True, 'ergonomicos_postursforzadas_h2': True,
                'quimicos_polvos_h1': True, 'quimicos_polvos_h2': True, 'quimicos_polvos_h3': True,
                'seguridad_mec_caidasadiferentenivel_h1': True,
                'psicosociales_sobrecargalaboral_h1': True, 'psicosociales_sobrecargalaboral_h2': True,
            },
            'medidas': 'USO OBLIGATORIO DE EPP: CASCO, GUANTES, BOTAS PUNTA DE ACERO, PROTECCIÓN AUDITIVA Y VISUAL\nPAUSAS ACTIVAS CADA 2 HORAS\nROTACIÓN DE ACTIVIDADES PARA EVITAR SOBRECARGA ERGONÓMICA\nCAPACITACIÓN EN MANEJO MANUAL DE CARGAS\nSEÑALIZACIÓN DE ZONAS DE RIESGO',
        },
        'h': {
            'empleos': [
                {'centro': 'CONSTRUCTORA MACHALA S.A.', 'cargo': 'OPERADOR MAQUINARIA', 'estado': 'ACTUAL', 'tiempo': '0 MESES (INGRESO)'},
                {'centro': 'MINERA GOLD S.A.', 'cargo': 'PERFORISTA', 'estado': 'ANTERIOR', 'tiempo': '48 MESES'},
                {'centro': 'TRANSPORTES DEL SUR', 'cargo': 'CONDUCTOR PESADO', 'estado': 'ANTERIOR', 'tiempo': '36 MESES'},
            ],
        },
        'i': {
            'actividades': [
                {'descripcion': 'FÚTBOL 2 VECES POR SEMANA'},
                {'descripcion': 'AGRICULTURA FAMILIAR EN FINCA LOS FINES DE SEMANA'},
            ],
        },
        'j': {
            'categorias': [
                {
                    'nombre': 'Hematología',
                    'fecha': '2026-02-28',
                    'items': [
                        {'nombre': 'Hemoglobina', 'valor': '15.2', 'unidad': 'g/dL'},
                        {'nombre': 'Hematocrito', 'valor': '45', 'unidad': '%'},
                        {'nombre': 'Leucocitos', 'valor': '7800', 'unidad': '/mm³'},
                        {'nombre': 'Plaquetas', 'valor': '250000', 'unidad': '/mm³'},
                    ]
                },
                {
                    'nombre': 'Química Sanguínea',
                    'fecha': '2026-02-28',
                    'items': [
                        {'nombre': 'Glucosa', 'valor': '95', 'unidad': 'mg/dL'},
                        {'nombre': 'Creatinina', 'valor': '0.9', 'unidad': 'mg/dL'},
                        {'nombre': 'Colesterol Total', 'valor': '210', 'unidad': 'mg/dL'},
                        {'nombre': 'Triglicéridos', 'valor': '165', 'unidad': 'mg/dL'},
                        {'nombre': 'Ácido úrico', 'valor': '6.2', 'unidad': 'mg/dL'},
                    ]
                },
                {
                    'nombre': 'Orina',
                    'fecha': '2026-02-28',
                    'items': [
                        {'nombre': 'EMO', 'valor': 'NORMAL', 'unidad': ''},
                    ]
                },
                {
                    'nombre': 'Imagenología',
                    'fecha': '2026-02-27',
                    'items': [
                        {'nombre': 'Rx Tórax', 'valor': 'NORMAL', 'unidad': ''},
                        {'nombre': 'Rx Columna Lumbar', 'valor': 'DISCOPATÍA DEGENERATIVA L4-L5 LEVE', 'unidad': ''},
                    ]
                },
                {
                    'nombre': 'Audiometría',
                    'fecha': '2026-02-27',
                    'items': [
                        {'nombre': 'Audiometría tonal', 'valor': 'HIPOACUSIA NEUROSENSORIAL LEVE BILATERAL EN FRECUENCIAS AGUDAS', 'unidad': ''},
                    ]
                },
                {
                    'nombre': 'Espirometría',
                    'fecha': '2026-02-27',
                    'items': [
                        {'nombre': 'Espirometría', 'valor': 'PATRÓN NORMAL - FVC 4.2L (95%) FEV1 3.5L (92%)', 'unidad': ''},
                    ]
                },
                {
                    'nombre': 'Optometría',
                    'fecha': '2026-02-27',
                    'items': [
                        {'nombre': 'Agudeza visual', 'valor': 'OD 20/20 OI 20/25 SC', 'unidad': ''},
                    ]
                },
            ],
            'observaciones': 'HALLAZGOS RELEVANTES: HIPOACUSIA NEUROSENSORIAL BILATERAL POR EXPOSICIÓN PREVIA A RUIDO. DISCOPATÍA L4-L5 LEVE.',
        },
        'k': {
            'diagnosticos': [
                {'codigo': 'Z00.0', 'descripcion': 'EXAMEN MÉDICO GENERAL', 'tipo': 'DEF'},
                {'codigo': 'Z57.1', 'descripcion': 'EXPOSICIÓN OCUPACIONAL A RUIDO', 'tipo': 'PRE'},
                {'codigo': 'H90.3', 'descripcion': 'HIPOACUSIA NEUROSENSORIAL BILATERAL', 'tipo': 'DEF'},
                {'codigo': 'M51.1', 'descripcion': 'DISCOPATÍA DEGENERATIVA LUMBAR', 'tipo': 'PRE'},
                {'codigo': 'E66.0', 'descripcion': 'SOBREPESO', 'tipo': 'DEF'},
                {'codigo': 'E78.5', 'descripcion': 'DISLIPIDEMIA MIXTA', 'tipo': 'PRE'},
            ]
        },
        'l': {
            'aptitud': 'APTO EN OBSERVACIÓN',
            'observaciones': 'PACIENTE AL MOMENTO EN BUENAS CONDICIONES DE SALUD, APTO EN OBSERVACIÓN PARA LABORAR EN EL PUESTO DE TRABAJO. REQUIERE CONTROL AUDIOMÉTRICO EN 6 MESES Y SEGUIMIENTO DE PATOLOGÍA LUMBAR.',
        },
        'm': {
            'estandar': [
                'ALIMENTACIÓN SALUDABLE HIPOGRASA, BAJA EN AZÚCARES Y HARINAS REFINADAS.',
                'ACTIVIDAD FÍSICA DE BAJO IMPACTO, AL MENOS 30 MINUTOS AL DÍA.',
                'HIDRATACIÓN CORPORAL ADECUADA.',
                'USO DE PRENDAS DE PREVENCIÓN Y CUMPLIMIENTO DE PROTOCOLOS DE BIOSEGURIDAD Y TRABAJO SEGURO.',
                'EVITAR O DISMINUIR MOVIMIENTOS FORZADOS Y/O REPETITIVOS.',
                'REALIZAR PAUSAS ACTIVAS DURANTE LA JORNADA LABORAL.',
                'EVITAR MOVIMIENTO MANUAL DE CARGA DE MANERA INCORRECTA Y CORREGIR MALAS ACTITUDES POSTURALES.',
                'ADOPTAR ACTITUDES ERGONÓMICAS ADECUADAS DE ACUERDO A LA NATURALEZA DEL CARGO.',
                'CONTROL MÉDICO OCUPACIONAL EN 6 MESES.',
                'EVITAR EXPOSICIÓN PROLONGADA A RUIDO.',
                'USO OBLIGATORIO DE EPP DURANTE LA JORNADA LABORAL.',
            ],
            'medicas': 'CONTROL AUDIOMÉTRICO SEMESTRAL. MANEJO CON TRAUMATOLOGÍA POR DISCOPATÍA LUMBAR. CONTROL METABÓLICO POR DISLIPIDEMIA.',
        },
        'n': {
            'medico': 'DR. LEONARDO ARIAS ESPINOZA',
            'codigo': 'MSP-12345',
        },
        'r': {
            'se_realiza': '',
            'condicion_trabajo': '',
            'observacion': '',
        },
    }
    
    result = fill_formulario(
        test_data,
        '/home/claude/formulario_final.xlsx',
        '/home/claude/test_llenado.pdf'
    )
    print(f"\nResultado: {result}")
